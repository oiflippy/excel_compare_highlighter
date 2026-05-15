from __future__ import annotations

import json
import os
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Any
import re

import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


APP_NAME = "Excel Compare Highlighter"
SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")
PREVIEW_LIMIT = 200
PROFILE_LIMIT = 30
BLANK_ROW_STOP_LIMIT = 200
SHEET_PREVIEW_ROWS = 30
SHEET_PREVIEW_COLUMNS = 40
COLOR_PATTERN = re.compile(r"^#[0-9a-fA-F]{6}$")


@dataclass
class SheetInfo:
    path: Path
    sheet: str
    header_row: int
    headers: list[str]


@dataclass
class RowRange:
    start: int
    end: int | None


@dataclass
class ColumnRange:
    start: int | None
    end: int | None


@dataclass
class CompareContext:
    source: SheetInfo
    source_paths: list[Path]
    target: SheetInfo
    source_column: int
    source_rule_column: int
    target_column: int
    source_rows: RowRange
    target_rows: RowRange
    highlight_columns: ColumnRange
    color_rules: dict[str, str]


def app_data_dir() -> Path:
    base = os.getenv("APPDATA") or str(Path.home())
    path = Path(base) / APP_NAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def profiles_path() -> Path:
    return app_data_dir() / "profiles.json"


def load_profiles() -> list[dict[str, Any]]:
    path = profiles_path()
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    return data if isinstance(data, list) else []


def save_profiles(profiles: list[dict[str, Any]]) -> None:
    profiles_path().write_text(
        json.dumps(profiles[:PROFILE_LIMIT], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any, *, case_sensitive: bool) -> str:
    text = str(value or "").strip()
    text = " ".join(text.split())
    return text if case_sensitive else text.casefold()


def validate_color_hex(color_hex: str) -> str:
    color = color_hex.strip()
    if not COLOR_PATTERN.match(color):
        raise ValueError("颜色必须是 #RRGGBB 格式，例如 #FFF2CC。")
    return color.upper()


def parse_optional_positive_int(value: str, field_name: str) -> int | None:
    text = value.strip()
    if not text:
        return None
    try:
        number = int(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} 必须是正整数，或留空表示自动。") from exc
    if number < 1:
        raise ValueError(f"{field_name} 必须大于 0。")
    return number


def parse_row_range(start_text: str, end_text: str, default_start: int, label: str) -> RowRange:
    start = parse_optional_positive_int(start_text, f"{label}起始行") or default_start
    end = parse_optional_positive_int(end_text, f"{label}结束行")
    if end is not None and end < start:
        raise ValueError(f"{label}结束行不能小于起始行。")
    return RowRange(start=start, end=end)


def parse_column_reference(value: str, field_name: str) -> int | None:
    text = value.strip()
    if not text:
        return None
    try:
        if text.isdigit():
            number = int(text)
        else:
            number = column_index_from_string(text.upper())
    except ValueError as exc:
        raise ValueError(f"{field_name} 必须是列号或列字母，例如 1 或 A；留空表示全选。") from exc
    if number < 1:
        raise ValueError(f"{field_name} 必须大于 0。")
    return number


def parse_column_range(start_text: str, end_text: str) -> ColumnRange:
    start = parse_column_reference(start_text, "填色开始列")
    end = parse_column_reference(end_text, "填色结束列")
    if (start is None) != (end is None):
        raise ValueError("填色开始列和结束列需要同时填写，或同时留空表示整行。")
    if start is not None and end is not None and end < start:
        raise ValueError("填色结束列不能小于开始列。")
    return ColumnRange(start=start, end=end)


def parse_color_rules(value: str, *, case_sensitive: bool) -> dict[str, str]:
    rules: dict[str, str] = {}
    text = value.strip()
    if not text:
        return rules
    for item in re.split(r"[;；,，\n]+", text):
        item = item.strip()
        if not item:
            continue
        if "=" not in item:
            raise ValueError("指定项颜色规则格式应为：A=#FFF2CC;B=#C6EFCE。")
        rule_value, color = item.split("=", 1)
        rule = normalize_key(rule_value, case_sensitive=case_sensitive)
        if not rule:
            raise ValueError("指定项颜色规则中存在空指定项。")
        rules[rule] = validate_color_hex(color)
    return rules


def ensure_supported_excel(path: Path) -> None:
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("当前版本支持 .xlsx / .xlsm / .xltx / .xltm，暂不支持旧版 .xls。")


def read_sheet_names(path: Path) -> list[str]:
    ensure_supported_excel(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_headers(path: Path, sheet_name: str, header_row: int) -> list[str]:
    ensure_supported_excel(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{path.name} 中没有工作表：{sheet_name}。")
        worksheet = workbook[sheet_name]

        max_col = worksheet.max_column or 1
        header_rows = worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
        header_values = list(next(header_rows, tuple()))
        headers = [normalize_header(cell) for cell in header_values]

        sample_values = first_data_row_values(worksheet, header_row + 1, max_col)
        for index, header in enumerate(headers):
            if header:
                continue
            sample = normalize_header(sample_values[index]) if index < len(sample_values) else ""
            if sample:
                shown = sample.replace("\n", " ")[:30]
                headers[index] = f"未命名列（示例：{shown}）"
            else:
                headers[index] = "未命名列"

        return headers
    finally:
        workbook.close()


def first_data_row_values(worksheet: Any, start_row: int, max_col: int) -> list[Any]:
    for _, row in read_effective_rows(
        worksheet,
        min_row=start_row,
        max_row=start_row + BLANK_ROW_STOP_LIMIT,
        min_col=1,
        max_col=max_col,
    ):
        if any(normalize_header(value) for value in row):
            return list(row)
    return []


def read_sheet_preview(
    path: Path,
    sheet_name: str,
    *,
    max_rows: int = SHEET_PREVIEW_ROWS,
    max_cols: int = SHEET_PREVIEW_COLUMNS,
) -> tuple[list[str], list[list[Any]]]:
    ensure_supported_excel(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{path.name} 中没有工作表：{sheet_name}。")
        worksheet = workbook[sheet_name]
        column_count = min(worksheet.max_column or 1, max_cols)
        columns = [get_column_letter(index) for index in range(1, column_count + 1)]
        rows: list[list[Any]] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=column_count,
            values_only=True,
            ),
            start=1,
        ):
            rows.append([row_number, *row])
        return columns, rows
    finally:
        workbook.close()


def column_options(headers: list[str]) -> list[str]:
    return [f"{get_column_letter(index)} | {header or '未命名列'}" for index, header in enumerate(headers, start=1)]


def parse_column_option(option: str) -> int:
    text = option.strip()
    if not text:
        raise ValueError("请选择或填写有效列。")
    if "|" in text:
        text = text.split("|", 1)[0].strip()
    try:
        if text.isdigit():
            return int(text)
        return column_index_from_string(text.upper())
    except ValueError as exc:
        raise ValueError("列必须是下拉选项、列字母或列号，例如 D 或 4。") from exc


def option_for_header(headers: list[str], wanted_header: str) -> str | None:
    wanted = normalize_header(wanted_header)
    for option, header in zip(column_options(headers), headers):
        if normalize_header(header) == wanted:
            return option
    wanted_folded = wanted.casefold()
    for option, header in zip(column_options(headers), headers):
        if normalize_header(header).casefold() == wanted_folded:
            return option
    return None


def header_signature(headers: list[str]) -> set[str]:
    return {normalize_header(header).casefold() for header in headers if normalize_header(header)}


def jaccard(left: set[str], right: set[str]) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / len(left | right)


def excel_files_in_source(path: Path) -> list[Path]:
    if path.is_dir():
        files = [item for item in sorted(path.iterdir()) if item.is_file()]
        excel_files = [item for item in files if item.suffix.lower() in SUPPORTED_EXTENSIONS and not item.name.startswith("~$")]
        if not excel_files:
            raise ValueError("表 A 文件夹中没有找到支持的 Excel 文件。")
        return excel_files
    ensure_supported_excel(path)
    return [path]


def validate_source_sheets(paths: list[Path], sheet_name: str, header_row: int | None = None) -> None:
    missing: list[str] = []
    for path in paths:
        if sheet_name not in read_sheet_names(path):
            missing.append(path.name)
    if missing:
        shown = "\n".join(f"- {name}" for name in missing[:20])
        more = "" if len(missing) <= 20 else f"\n... 还有 {len(missing) - 20} 个文件"
        raise ValueError(f"表 A 文件夹中这些文件没有 Sheet“{sheet_name}”：\n{shown}{more}")
    if header_row is None or len(paths) <= 1:
        return

    reference_path = paths[0]
    reference_headers = read_headers(reference_path, sheet_name, header_row)
    mismatched: list[str] = []
    for path in paths[1:]:
        headers = read_headers(path, sheet_name, header_row)
        if headers != reference_headers:
            mismatched.append(path.name)
    if mismatched:
        shown = "\n".join(f"- {name}" for name in mismatched[:20])
        more = "" if len(mismatched) <= 20 else f"\n... 还有 {len(mismatched) - 20} 个文件"
        raise ValueError(
            f"表 A 文件夹中这些文件的表头/列结构与 {reference_path.name} 不一致：\n"
            f"{shown}{more}\n\n请确保文件夹中的 A 表格式一样，只允许数据内容不同。"
        )


def read_effective_rows(
    worksheet: Any,
    *,
    min_row: int,
    max_row: int | None,
    min_col: int,
    max_col: int,
) -> Any:
    blank_rows = 0
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=min_row,
    ):
        if any(cell not in (None, "") for cell in row):
            blank_rows = 0
            yield row_number, row
            continue
        blank_rows += 1
        if max_row is None and blank_rows >= BLANK_ROW_STOP_LIMIT:
            break


def build_key_set(
    sheet: SheetInfo,
    column_index: int,
    row_range: RowRange,
    *,
    case_sensitive: bool,
) -> tuple[set[str], int, int]:
    workbook = load_workbook(sheet.path, read_only=True, data_only=True)
    keys: set[str] = set()
    total_rows = 0
    blank_rows = 0
    try:
        worksheet = workbook[sheet.sheet]
        for row in worksheet.iter_rows(
            min_row=row_range.start,
            max_row=row_range.end,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            total_rows += 1
            value = row[0] if row else None
            key = normalize_key(value, case_sensitive=case_sensitive)
            if key:
                keys.add(key)
            else:
                blank_rows += 1
        return keys, total_rows, blank_rows
    finally:
        workbook.close()


def build_rule_map(
    sheets: list[SheetInfo],
    match_column: int,
    rule_column: int,
    row_range: RowRange,
    *,
    case_sensitive: bool,
) -> tuple[dict[str, set[str]], int, int]:
    rule_map: dict[str, set[str]] = {}
    total_rows = 0
    blank_rows = 0
    for sheet in sheets:
        workbook = load_workbook(sheet.path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet.sheet]
            max_column = max(match_column, rule_column)
            for _, row in read_effective_rows(
                worksheet,
                min_row=row_range.start,
                max_row=row_range.end,
                min_col=1,
                max_col=max_column,
            ):
                total_rows += 1
                match_value = row[match_column - 1] if len(row) >= match_column else None
                rule_value = row[rule_column - 1] if len(row) >= rule_column else None
                key = normalize_key(match_value, case_sensitive=case_sensitive)
                rule = normalize_key(rule_value, case_sensitive=case_sensitive)
                if key:
                    rule_map.setdefault(key, set())
                    if rule:
                        rule_map[key].add(rule)
                else:
                    blank_rows += 1
        finally:
            workbook.close()
    return rule_map, total_rows, blank_rows


def scan_target_matches(
    sheet: SheetInfo,
    column_index: int,
    row_range: RowRange,
    keys: set[str],
    *,
    case_sensitive: bool,
    preview_limit: int | None = PREVIEW_LIMIT,
) -> tuple[list[int], list[list[Any]], int]:
    workbook = load_workbook(sheet.path, read_only=True, data_only=True)
    matched_rows: list[int] = []
    preview_rows: list[list[Any]] = []
    scanned = 0
    try:
        worksheet = workbook[sheet.sheet]
        max_column = max(column_index, 8 if preview_limit else column_index)
        for row_number, row in read_effective_rows(
            worksheet,
            min_row=row_range.start,
            max_row=row_range.end,
            min_col=1,
            max_col=max_column,
        ):
            scanned += 1
            value = row[column_index - 1] if len(row) >= column_index else None
            key = normalize_key(value, case_sensitive=case_sensitive)
            matched = bool(key and key in keys)
            if matched:
                matched_rows.append(row_number)
                if preview_limit is None or len(preview_rows) < preview_limit:
                    preview_rows.append([row_number, "", *list(row[: min(len(row), 8)])])
        return matched_rows, preview_rows, scanned
    finally:
        workbook.close()


def scan_target_rule_matches(
    sheet: SheetInfo,
    column_index: int,
    row_range: RowRange,
    rule_map: dict[str, set[str]],
    color_rules: dict[str, str],
    *,
    case_sensitive: bool,
    preview_limit: int | None = PREVIEW_LIMIT,
) -> tuple[dict[int, str], list[list[Any]], int]:
    workbook = load_workbook(sheet.path, read_only=True, data_only=True)
    matched_rows: dict[int, str] = {}
    preview_rows: list[list[Any]] = []
    scanned = 0
    try:
        worksheet = workbook[sheet.sheet]
        max_column = max(column_index, 8 if preview_limit else column_index)
        for row_number, row in read_effective_rows(
            worksheet,
            min_row=row_range.start,
            max_row=row_range.end,
            min_col=1,
            max_col=max_column,
        ):
            scanned += 1
            value = row[column_index - 1] if len(row) >= column_index else None
            key = normalize_key(value, case_sensitive=case_sensitive)
            rules = rule_map.get(key, set()) if key else set()
            if not rules:
                continue
            color = ""
            for rule, rule_color in color_rules.items():
                if rule in rules:
                    color = rule_color
                    break
            if not color:
                continue
            matched_rows[row_number] = color
            if preview_limit is None or len(preview_rows) < preview_limit:
                preview_rows.append([row_number, color, *list(row[: min(len(row), 8)])])
        return matched_rows, preview_rows, scanned
    finally:
        workbook.close()


def apply_highlight(
    target: SheetInfo,
    matched_rows: list[int] | dict[int, str],
    color_hex: str,
    column_range: ColumnRange,
    output_path: Path,
) -> None:
    keep_vba = target.path.suffix.lower() in {".xlsm", ".xltm"}
    workbook = load_workbook(target.path, keep_vba=keep_vba)
    try:
        worksheet = workbook[target.sheet]
        row_colors = matched_rows if isinstance(matched_rows, dict) else {row_number: color_hex for row_number in matched_rows}
        for row_number, row_color in row_colors.items():
            color = validate_color_hex(row_color).replace("#", "")
            fill = PatternFill(fill_type="solid", fgColor=color)
            start_column = column_range.start or 1
            end_column = column_range.end or worksheet.max_column
            for row in worksheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                min_col=start_column,
                max_col=end_column,
            ):
                for cell in row:
                    cell.fill = fill
        workbook.save(output_path)
    finally:
        workbook.close()


class ExcelCompareApp(ttk.Frame):
    def __init__(self, master: tk.Tk) -> None:
        super().__init__(master, padding=12)
        self.master = master
        self.grid(sticky="nsew")
        master.title("Excel 表格对比高亮工具")
        master.geometry("980x720")
        master.minsize(900, 640)
        master.columnconfigure(0, weight=1)
        master.rowconfigure(0, weight=1)

        self.source_path = tk.StringVar()
        self.target_path = tk.StringVar()
        self.source_sheet = tk.StringVar()
        self.target_sheet = tk.StringVar()
        self.source_header_row = tk.IntVar(value=1)
        self.target_header_row = tk.IntVar(value=1)
        self.source_column = tk.StringVar()
        self.source_rule_column = tk.StringVar()
        self.target_column = tk.StringVar()
        self.source_start_row = tk.StringVar()
        self.source_end_row = tk.StringVar()
        self.target_start_row = tk.StringVar()
        self.target_end_row = tk.StringVar()
        self.highlight_start_column = tk.StringVar()
        self.highlight_end_column = tk.StringVar()
        self.rule_text = tk.StringVar()
        self.case_sensitive = tk.BooleanVar(value=False)
        self.status = tk.StringVar(value="请选择表 A 数据源和表 B 待填色文件。")

        self.source_headers: list[str] = []
        self.target_headers: list[str] = []
        self.latest_matches: list[int] = []
        self.latest_keys_count = 0
        self._busy = False
        self.task_buttons: list[ttk.Widget] = []

        self._build_ui()

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        file_frame = ttk.LabelFrame(self, text="1. 选择文件", padding=10)
        file_frame.grid(row=0, column=0, sticky="ew")
        file_frame.columnconfigure(1, weight=1)

        self.task_buttons.append(self._file_row(file_frame, 0, "表 A：来源表/文件夹", self.source_path, self.pick_source_file))
        self.task_buttons.append(self._file_row(file_frame, 1, "表 B：要填色的表", self.target_path, self.pick_target_file))

        config_frame = ttk.LabelFrame(self, text="2. 设置流程：A 表指定项 + A 表查找项 → B 表查找项 → 按规则给 B 行填色", padding=10)
        config_frame.grid(row=1, column=0, sticky="ew", pady=(10, 10))
        for index in (1, 3, 5):
            config_frame.columnconfigure(index, weight=1)

        ttk.Label(config_frame, text="表 A Sheet").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        self.source_sheet_box = ttk.Combobox(config_frame, textvariable=self.source_sheet, state="readonly")
        self.source_sheet_box.grid(row=0, column=1, sticky="ew", pady=4)

        ttk.Label(config_frame, text="表 A 表头行").grid(row=0, column=2, sticky="w", padx=(12, 6), pady=4)
        ttk.Spinbox(config_frame, from_=1, to=100, textvariable=self.source_header_row, width=8).grid(row=0, column=3, sticky="w", pady=4)

        self.reload_source_button = ttk.Button(config_frame, text="刷新 A 表", command=self.reload_source_columns)
        self.reload_source_button.grid(row=0, column=4, sticky="w", padx=(12, 0), pady=4)
        self.task_buttons.append(self.reload_source_button)

        ttk.Label(config_frame, text="表 B Sheet").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=4)
        self.target_sheet_box = ttk.Combobox(config_frame, textvariable=self.target_sheet, state="readonly")
        self.target_sheet_box.grid(row=1, column=1, sticky="ew", pady=4)

        ttk.Label(config_frame, text="表 B 表头行").grid(row=1, column=2, sticky="w", padx=(12, 6), pady=4)
        ttk.Spinbox(config_frame, from_=1, to=100, textvariable=self.target_header_row, width=8).grid(row=1, column=3, sticky="w", pady=4)

        self.reload_target_button = ttk.Button(config_frame, text="刷新 B 表", command=self.reload_target_columns)
        self.reload_target_button.grid(row=1, column=4, sticky="w", padx=(12, 0), pady=4)
        self.task_buttons.append(self.reload_target_button)

        ttk.Label(config_frame, text="A 指定项列").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=4)
        source_rule_frame = ttk.Frame(config_frame)
        source_rule_frame.grid(row=2, column=1, sticky="ew", pady=4)
        source_rule_frame.columnconfigure(0, weight=1)
        self.source_rule_column_entry = ttk.Entry(source_rule_frame, textvariable=self.source_rule_column, state="readonly")
        self.source_rule_column_entry.grid(row=0, column=0, sticky="ew")
        self.source_rule_pick_button = ttk.Button(source_rule_frame, text="选列", command=lambda: self.pick_column_from_sheet("source_rule"), width=6)
        self.source_rule_pick_button.grid(row=0, column=1, padx=(6, 0))

        ttk.Label(config_frame, text="A 查找项列").grid(row=2, column=2, sticky="w", padx=(12, 6), pady=4)
        source_column_frame = ttk.Frame(config_frame)
        source_column_frame.grid(row=2, column=3, sticky="ew", pady=4)
        source_column_frame.columnconfigure(0, weight=1)
        self.source_column_entry = ttk.Entry(source_column_frame, textvariable=self.source_column, state="readonly")
        self.source_column_entry.grid(row=0, column=0, sticky="ew")
        self.source_pick_button = ttk.Button(source_column_frame, text="选列", command=lambda: self.pick_column_from_sheet("source"), width=6)
        self.source_pick_button.grid(row=0, column=1, padx=(6, 0))

        ttk.Label(config_frame, text="B 查找项列").grid(row=2, column=4, sticky="w", padx=(12, 6), pady=4)
        target_column_frame = ttk.Frame(config_frame)
        target_column_frame.grid(row=2, column=5, sticky="ew", pady=4)
        target_column_frame.columnconfigure(0, weight=1)
        self.target_column_entry = ttk.Entry(target_column_frame, textvariable=self.target_column, state="readonly")
        self.target_column_entry.grid(row=0, column=0, sticky="ew")
        self.target_pick_button = ttk.Button(target_column_frame, text="选列", command=lambda: self.pick_column_from_sheet("target"), width=6)
        self.target_pick_button.grid(row=0, column=1, padx=(6, 0))
        self.task_buttons.extend([self.source_rule_pick_button, self.source_pick_button, self.target_pick_button])

        ttk.Label(config_frame, text="A 数据行").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=4)
        self._range_picker(config_frame, 3, 1, self.source_start_row, self.source_end_row, "默认表头下一行到末尾", lambda: self.pick_row_range_from_sheet("source"))

        ttk.Label(config_frame, text="B 查找行").grid(row=3, column=2, sticky="w", padx=(12, 6), pady=4)
        self._range_picker(config_frame, 3, 3, self.target_start_row, self.target_end_row, "默认表头下一行到末尾", lambda: self.pick_row_range_from_sheet("target"))

        ttk.Label(config_frame, text="填色列").grid(row=3, column=4, sticky="w", padx=(12, 6), pady=4)
        self._range_picker(config_frame, 3, 5, self.highlight_start_column, self.highlight_end_column, "留空为整行", self.pick_highlight_columns_from_sheet)

        ttk.Label(config_frame, text="指定项规则").grid(row=4, column=0, sticky="w", padx=(0, 6), pady=4)
        ttk.Entry(config_frame, textvariable=self.rule_text, state="readonly").grid(row=4, column=1, columnspan=4, sticky="ew", pady=4)
        self.rule_button = ttk.Button(config_frame, text="配置颜色", command=self.configure_color_rules)
        self.rule_button.grid(row=4, column=5, sticky="ew", padx=(12, 0), pady=4)
        self.task_buttons.append(self.rule_button)

        options = ttk.Frame(config_frame)
        options.grid(row=5, column=0, columnspan=3, sticky="w", pady=(8, 0))
        ttk.Checkbutton(options, text="区分大小写", variable=self.case_sensitive).pack(side="left")
        ttk.Label(options, text="只给已配置颜色的指定项填色；没选颜色的行保持原样", foreground="#666666").pack(side="left", padx=(16, 0))

        actions = ttk.Frame(config_frame)
        actions.grid(row=5, column=3, columnspan=3, sticky="e", pady=(8, 0))
        self.preview_button = ttk.Button(actions, text="预览", command=self.preview)
        self.preview_button.pack(side="left", padx=(0, 8))
        self.export_button = ttk.Button(actions, text="生成高亮文件", command=self.export)
        self.export_button.pack(side="left")
        self.task_buttons.extend([self.preview_button, self.export_button])

        preview_frame = ttk.LabelFrame(self, text="3. 快速预览命中行（最多显示 200 行，每行显示颜色和前 8 列）", padding=10)
        preview_frame.grid(row=2, column=0, sticky="nsew")
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)

        columns = ["row", "color", *[f"c{i}" for i in range(1, 9)]]
        self.preview_table = ttk.Treeview(preview_frame, columns=columns, show="headings", height=18)
        self.preview_table.heading("row", text="Excel 行号")
        self.preview_table.column("row", width=90, anchor="center")
        self.preview_table.heading("color", text="填充颜色")
        self.preview_table.column("color", width=100, anchor="center")
        for index in range(1, 9):
            column = f"c{index}"
            self.preview_table.heading(column, text=f"第 {index} 列")
            self.preview_table.column(column, width=120, anchor="w")
        y_scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_table.yview)
        x_scroll = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.preview_table.xview)
        self.preview_table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview_table.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        status_frame = ttk.Frame(self)
        status_frame.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        status_frame.columnconfigure(0, weight=1)
        ttk.Label(status_frame, textvariable=self.status).grid(row=0, column=0, sticky="w")

        self.source_sheet_box.bind("<<ComboboxSelected>>", lambda _event: self.reload_source_columns(silent=True))
        self.target_sheet_box.bind("<<ComboboxSelected>>", lambda _event: self.reload_target_columns(silent=True))

    def _file_row(self, parent: ttk.Frame, row: int, label: str, variable: tk.StringVar, command: Any) -> ttk.Button:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", pady=4)
        button = ttk.Button(parent, text="浏览...", command=command)
        button.grid(row=row, column=2, sticky="e", padx=(8, 0), pady=4)
        return button

    def _range_inputs(
        self,
        parent: ttk.Frame,
        row: int,
        column: int,
        start_var: tk.StringVar,
        end_var: tk.StringVar,
        tooltip: str,
    ) -> None:
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=column, sticky="ew", pady=4)
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(2, weight=1)
        ttk.Entry(frame, textvariable=start_var, width=8).grid(row=0, column=0, sticky="ew")
        ttk.Label(frame, text="到").grid(row=0, column=1, padx=4)
        ttk.Entry(frame, textvariable=end_var, width=8).grid(row=0, column=2, sticky="ew")
        ttk.Label(frame, text=tooltip, foreground="#666666").grid(row=1, column=0, columnspan=3, sticky="w")

    def _range_picker(
        self,
        parent: ttk.Frame,
        row: int,
        column: int,
        start_var: tk.StringVar,
        end_var: tk.StringVar,
        tooltip: str,
        command: Any,
    ) -> None:
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=column, sticky="ew", pady=4)
        frame.columnconfigure(0, weight=1)
        display_var = tk.StringVar()

        def refresh_display(*_: Any) -> None:
            start = start_var.get().strip()
            end = end_var.get().strip()
            if start and end:
                display_var.set(f"{start} 到 {end}")
            elif start:
                display_var.set(f"从 {start} 到末尾")
            else:
                display_var.set("")

        start_var.trace_add("write", refresh_display)
        end_var.trace_add("write", refresh_display)
        refresh_display()
        ttk.Entry(frame, textvariable=display_var, state="readonly", width=12).grid(row=0, column=0, sticky="ew")
        button = ttk.Button(frame, text="预览选择", command=command, width=9)
        button.grid(row=0, column=1, padx=(6, 0))
        ttk.Label(frame, text=tooltip, foreground="#666666").grid(row=1, column=0, columnspan=2, sticky="w")
        self.task_buttons.append(button)

    def pick_source_file(self) -> None:
        choice = messagebox.askyesno("选择表 A", "表 A 是否选择整个文件夹？\n\n是：选择文件夹\n否：选择单个 Excel 文件")
        if choice:
            folder = filedialog.askdirectory(title="选择表 A 文件夹")
            if not folder:
                return
            self._clear_source_columns()
            self.source_path.set(folder)
            self._run_task(lambda: self._load_sheets(excel_files_in_source(Path(folder))[0], self.source_sheet_box, self.source_sheet))
        else:
            self._clear_source_columns()
            self._pick_file(self.source_path, self.source_sheet_box, self.source_sheet)

    def pick_target_file(self) -> None:
        self._clear_target_columns()
        self._pick_file(self.target_path, self.target_sheet_box, self.target_sheet)

    def _clear_source_columns(self) -> None:
        self.source_headers = []
        self.source_column.set("")
        self.source_rule_column.set("")

    def _clear_target_columns(self) -> None:
        self.target_headers = []
        self.target_column.set("")

    def _pick_file(self, path_var: tk.StringVar, sheet_box: ttk.Combobox, sheet_var: tk.StringVar) -> None:
        filename = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")],
        )
        if not filename:
            return
        path_var.set(filename)
        self._run_task(lambda: self._load_sheets(Path(filename), sheet_box, sheet_var))

    def _load_sheets(self, path: Path, sheet_box: ttk.Combobox, sheet_var: tk.StringVar) -> None:
        names = read_sheet_names(path)
        self.master.after(0, lambda: self._set_sheets(names, sheet_box, sheet_var))

    def _set_sheets(self, names: list[str], sheet_box: ttk.Combobox, sheet_var: tk.StringVar) -> None:
        sheet_box.configure(values=names)
        if names:
            sheet_var.set(names[0])
        self.status.set("已读取工作表，请分别读取 A/B 列名。")

    def reload_columns(self, silent: bool = False) -> None:
        source: SheetInfo | None = None
        target: SheetInfo | None = None
        errors: list[str] = []
        try:
            source = self._sheet_info(kind="source", read_headers_now=True)
        except Exception as exc:
            errors.append(f"表 A：{exc}")
        try:
            target = self._sheet_info(kind="target", read_headers_now=True)
        except Exception as exc:
            errors.append(f"表 B：{exc}")
        if source:
            self._apply_source_headers(source)
        if target:
            self._apply_target_headers(target)
        if source and target:
            profile_name = self.apply_best_profile(source, target)
            if profile_name:
                self.status.set(f"已自动套用相似配置：{profile_name}")
            else:
                self.status.set("A/B 列名读取完成，请选择列后预览。")
        else:
            message = "；".join(errors) or "请先选择表 A 和表 B。"
            self.status.set(message)
            if not silent:
                messagebox.showerror("读取列名失败", message)

    def reload_source_columns(self, silent: bool = False) -> None:
        try:
            source = self._sheet_info(kind="source", read_headers_now=True)
        except Exception as exc:
            self.status.set(f"表 A 列名读取失败：{exc}")
            if not silent:
                messagebox.showerror("读取 A 列名失败", str(exc))
            return
        self._apply_source_headers(source)
        self.status.set("表 A 已刷新：请点击选列，在表格里点选指定项列和查找项列。")
        self._apply_profile_if_ready()

    def reload_target_columns(self, silent: bool = False) -> None:
        try:
            target = self._sheet_info(kind="target", read_headers_now=True)
        except Exception as exc:
            self.status.set(f"表 B 列名读取失败：{exc}")
            if not silent:
                messagebox.showerror("读取 B 列名失败", str(exc))
            return
        self._apply_target_headers(target)
        self.status.set("表 B 已刷新：请点击选列，在表格里点选 B 查找项列。")
        self._apply_profile_if_ready()

    def _apply_source_headers(self, source: SheetInfo) -> None:
        self.source_headers = source.headers
        options = column_options(source.headers)
        if options and not self.source_column.get():
            self.source_column.set(options[0])
        if options and not self.source_rule_column.get():
            self.source_rule_column.set(options[0])

    def _apply_target_headers(self, target: SheetInfo) -> None:
        self.target_headers = target.headers
        options = column_options(target.headers)
        if options and not self.target_column.get():
            self.target_column.set(options[0])

    def _source_preview_file(self) -> tuple[Path, list[Path]]:
        if not self.source_path.get():
            raise ValueError("请先选择表 A。")
        paths = excel_files_in_source(Path(self.source_path.get()))
        return paths[0], paths

    def pick_column_from_sheet(self, target: str) -> None:
        try:
            path, sheet, variable, headers, title = self._column_picker_context(target)
            columns, rows = read_sheet_preview(path, sheet)
        except Exception as exc:
            self._show_error(exc)
            return
        self._show_column_picker_dialog(title, columns, rows, variable, headers)

    def _column_picker_context(self, target: str) -> tuple[Path, str, tk.StringVar, list[str], str]:
        if target in {"source", "source_rule"}:
            path_text = self.source_path.get()
            sheet = self.source_sheet.get()
            header_row = self.source_header_row.get()
            variable = self.source_rule_column if target == "source_rule" else self.source_column
            title = "选择 A 指定项列" if target == "source_rule" else "选择 A 查找项列"
            path, paths = self._source_preview_file()
            validate_source_sheets(paths, sheet, header_row)
            self.source_headers = read_headers(path, sheet, header_row)
            if len(paths) > 1:
                title = f"{title}（文件夹模式，预览：{path.name}）"
            return path, sheet, variable, self.source_headers, title

        path_text = self.target_path.get()
        sheet = self.target_sheet.get()
        header_row = self.target_header_row.get()
        if not path_text:
            raise ValueError("请先选择表 B。")
        path = Path(path_text)
        self.target_headers = read_headers(path, sheet, header_row)
        return path, sheet, self.target_column, self.target_headers, "选择 B 查找项列"

    def _show_column_picker_dialog(
        self,
        title: str,
        columns: list[str],
        rows: list[list[Any]],
        variable: tk.StringVar,
        headers: list[str],
    ) -> None:
        dialog = tk.Toplevel(self.master)
        dialog.title(title)
        dialog.geometry("1100x620")
        dialog.transient(self.master)
        dialog.grab_set()

        ttk.Label(
            dialog,
            text="看表格内容选择列：点击任意单元格会选中它所在的列，确认后回填。",
            padding=10,
        ).pack(fill="x")

        table_frame = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        table_frame.pack(fill="both", expand=True)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        tree_columns = ["row", *columns]
        tree = ttk.Treeview(table_frame, columns=tree_columns, show="headings", height=22)
        tree.heading("row", text="行号")
        tree.column("row", width=56, minwidth=48, anchor="center", stretch=False)
        for index, column_letter in enumerate(columns, start=1):
            header = headers[index - 1] if index <= len(headers) else ""
            heading = column_letter if not header else f"{column_letter}｜{header[:12]}"
            tree.heading(column_letter, text=heading)
            tree.column(column_letter, width=130, minwidth=80, anchor="w")

        for row in rows:
            values = ["" if value is None else str(value).replace("\n", " ") for value in row]
            tree.insert("", "end", values=values)

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        selected_column = [self._current_column_index(variable.get())]
        selected_text = tk.StringVar(value=self._column_choice_text(selected_column[0], headers))

        def select_column(column_id: str) -> None:
            if column_id == "row" or column_id not in columns:
                return
            selected_column[0] = column_index_from_string(column_id)
            selected_text.set(self._column_choice_text(selected_column[0], headers))

        def on_click(event: tk.Event) -> None:
            column_id = tree.identify_column(event.x)
            if not column_id:
                return
            index = int(column_id.replace("#", "")) - 2
            if 0 <= index < len(columns):
                select_column(columns[index])

        tree.bind("<ButtonRelease-1>", on_click)

        button_frame = ttk.Frame(dialog, padding=10)
        button_frame.pack(fill="x")
        ttk.Label(button_frame, textvariable=selected_text).pack(side="left")

        def confirm() -> None:
            if selected_column[0] is None:
                messagebox.showwarning("请选择列", "请先点击表格中的某一列。", parent=dialog)
                return
            variable.set(self._column_choice_text(selected_column[0], headers))
            dialog.destroy()

        ttk.Button(button_frame, text="确认", command=confirm, width=12).pack(side="right", padx=(6, 0))
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

    def _current_column_index(self, option: str) -> int | None:
        try:
            return parse_column_option(option)
        except ValueError:
            return None

    def _column_choice_text(self, column_index: int | None, headers: list[str]) -> str:
        if column_index is None:
            return "未选择"
        header = headers[column_index - 1] if column_index <= len(headers) else ""
        return f"{get_column_letter(column_index)} | {header or '未命名列'}"

    def pick_row_range_from_sheet(self, kind: str) -> None:
        try:
            path, sheet, start_var, end_var, title, default_start = self._row_range_picker_context(kind)
            columns, rows = read_sheet_preview(path, sheet)
        except Exception as exc:
            self._show_error(exc)
            return
        self._show_row_range_dialog(title, columns, rows, start_var, end_var, default_start)

    def _row_range_picker_context(self, kind: str) -> tuple[Path, str, tk.StringVar, tk.StringVar, str, int]:
        if kind == "source":
            path, paths = self._source_preview_file()
            sheet = self.source_sheet.get()
            validate_source_sheets(paths, sheet, self.source_header_row.get())
            title = "选择 A 数据行"
            if len(paths) > 1:
                title = f"{title}（文件夹模式，预览：{path.name}）"
            return path, sheet, self.source_start_row, self.source_end_row, title, self.source_header_row.get() + 1
        if not self.target_path.get():
            raise ValueError("请先选择表 B。")
        return Path(self.target_path.get()), self.target_sheet.get(), self.target_start_row, self.target_end_row, "选择 B 查找行", self.target_header_row.get() + 1

    def _show_row_range_dialog(
        self,
        title: str,
        columns: list[str],
        rows: list[list[Any]],
        start_var: tk.StringVar,
        end_var: tk.StringVar,
        default_start: int,
    ) -> None:
        dialog = tk.Toplevel(self.master)
        dialog.title(title)
        dialog.geometry("1100x620")
        dialog.transient(self.master)
        dialog.grab_set()
        ttk.Label(dialog, text="点击第一行和最后一行；只点第一行表示从该行到末尾。", padding=10).pack(fill="x")

        table_frame = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        table_frame.pack(fill="both", expand=True)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        tree_columns = ["row", *columns]
        tree = ttk.Treeview(table_frame, columns=tree_columns, show="headings", height=22, selectmode="browse")
        tree.heading("row", text="行号")
        tree.column("row", width=56, minwidth=48, anchor="center", stretch=False)
        for column_letter in columns:
            tree.heading(column_letter, text=column_letter)
            tree.column(column_letter, width=130, minwidth=80, anchor="w")
        for row in rows:
            values = ["" if value is None else str(value).replace("\n", " ") for value in row]
            tree.insert("", "end", values=values)
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        selected: list[int] = []
        selected_text = tk.StringVar(value=f"默认：从第 {default_start} 行到末尾")

        def update_text() -> None:
            if not selected:
                selected_text.set(f"默认：从第 {default_start} 行到末尾")
            elif len(selected) == 1:
                selected_text.set(f"已选：从第 {selected[0]} 行到末尾")
            else:
                start, end = sorted(selected[:2])
                selected_text.set(f"已选：第 {start} 行到第 {end} 行")

        def on_select(_: tk.Event) -> None:
            selection = tree.selection()
            if not selection:
                return
            values = tree.item(selection[0], "values")
            row_number = int(values[0])
            if len(selected) >= 2:
                selected.clear()
            selected.append(row_number)
            update_text()

        tree.bind("<<TreeviewSelect>>", on_select)
        button_frame = ttk.Frame(dialog, padding=10)
        button_frame.pack(fill="x")
        ttk.Label(button_frame, textvariable=selected_text).pack(side="left")

        def use_default() -> None:
            start_var.set("")
            end_var.set("")
            dialog.destroy()

        def confirm() -> None:
            if not selected:
                start_var.set("")
                end_var.set("")
            elif len(selected) == 1:
                start_var.set(str(selected[0]))
                end_var.set("")
            else:
                start, end = sorted(selected[:2])
                start_var.set(str(start))
                end_var.set(str(end))
            dialog.destroy()

        ttk.Button(button_frame, text="使用默认", command=use_default, width=12).pack(side="right", padx=(6, 0))
        ttk.Button(button_frame, text="确认", command=confirm, width=12).pack(side="right", padx=(6, 0))
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

    def pick_highlight_columns_from_sheet(self) -> None:
        try:
            if not self.target_path.get():
                raise ValueError("请先选择表 B。")
            path = Path(self.target_path.get())
            sheet = self.target_sheet.get()
            columns, rows = read_sheet_preview(path, sheet)
            headers = read_headers(path, sheet, self.target_header_row.get())
        except Exception as exc:
            self._show_error(exc)
            return
        self._show_column_range_dialog("选择 B 填色列", columns, rows, self.highlight_start_column, self.highlight_end_column, headers)

    def _show_column_range_dialog(
        self,
        title: str,
        columns: list[str],
        rows: list[list[Any]],
        start_var: tk.StringVar,
        end_var: tk.StringVar,
        headers: list[str],
    ) -> None:
        dialog = tk.Toplevel(self.master)
        dialog.title(title)
        dialog.geometry("1100x620")
        dialog.transient(self.master)
        dialog.grab_set()
        ttk.Label(dialog, text="点击第一列和最后一列；点“整行填色”表示命中行整行填色。", padding=10).pack(fill="x")
        table_frame = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        table_frame.pack(fill="both", expand=True)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        tree_columns = ["row", *columns]
        tree = ttk.Treeview(table_frame, columns=tree_columns, show="headings", height=22)
        tree.heading("row", text="行号")
        tree.column("row", width=56, minwidth=48, anchor="center", stretch=False)
        for index, column_letter in enumerate(columns, start=1):
            header = headers[index - 1] if index <= len(headers) else ""
            tree.heading(column_letter, text=column_letter if not header else f"{column_letter}｜{header[:12]}")
            tree.column(column_letter, width=130, minwidth=80, anchor="w")
        for row in rows:
            values = ["" if value is None else str(value).replace("\n", " ") for value in row]
            tree.insert("", "end", values=values)
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        selected: list[int] = []
        selected_text = tk.StringVar(value="默认：整行填色")

        def update_text() -> None:
            if not selected:
                selected_text.set("默认：整行填色")
            elif len(selected) == 1:
                selected_text.set(f"已选：{get_column_letter(selected[0])} 列")
            else:
                start, end = sorted(selected[:2])
                selected_text.set(f"已选：{get_column_letter(start)} 到 {get_column_letter(end)} 列")

        def on_click(event: tk.Event) -> None:
            column_id = tree.identify_column(event.x)
            if not column_id:
                return
            index = int(column_id.replace("#", "")) - 2
            if not (0 <= index < len(columns)):
                return
            if len(selected) >= 2:
                selected.clear()
            selected.append(column_index_from_string(columns[index]))
            update_text()

        tree.bind("<ButtonRelease-1>", on_click)
        button_frame = ttk.Frame(dialog, padding=10)
        button_frame.pack(fill="x")
        ttk.Label(button_frame, textvariable=selected_text).pack(side="left")

        def use_full_row() -> None:
            start_var.set("")
            end_var.set("")
            dialog.destroy()

        def confirm() -> None:
            if not selected:
                start_var.set("")
                end_var.set("")
            elif len(selected) == 1:
                letter = get_column_letter(selected[0])
                start_var.set(letter)
                end_var.set(letter)
            else:
                start, end = sorted(selected[:2])
                start_var.set(get_column_letter(start))
                end_var.set(get_column_letter(end))
            dialog.destroy()

        ttk.Button(button_frame, text="整行填色", command=use_full_row, width=12).pack(side="right", padx=(6, 0))
        ttk.Button(button_frame, text="确认", command=confirm, width=12).pack(side="right", padx=(6, 0))
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

    def configure_color_rules(self) -> None:
        try:
            values = self._source_rule_values()
        except Exception as exc:
            self._show_error(exc)
            return
        self._show_rule_color_dialog(values)

    def _source_rule_values(self) -> list[str]:
        source = self._sheet_info(kind="source", read_headers_now=not self.source_headers)
        source_paths = excel_files_in_source(Path(self.source_path.get()))
        validate_source_sheets(source_paths, source.sheet, source.header_row)
        rule_column = parse_column_option(self.source_rule_column.get())
        row_range = parse_row_range(self.source_start_row.get(), self.source_end_row.get(), source.header_row + 1, "表 A ")
        values: set[str] = set()
        for path in source_paths:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                worksheet = workbook[source.sheet]
                for _, row in read_effective_rows(worksheet, min_row=row_range.start, max_row=row_range.end, min_col=1, max_col=rule_column):
                    value = normalize_header(row[rule_column - 1] if len(row) >= rule_column else None)
                    if value:
                        values.add(value)
            finally:
                workbook.close()
        if not values:
            raise ValueError("A 指定项列没有读取到可配置颜色的值。")
        return sorted(values)

    def _show_rule_color_dialog(self, values: list[str]) -> None:
        dialog = tk.Toplevel(self.master)
        dialog.title("配置指定项颜色")
        dialog.geometry("520x520")
        dialog.transient(self.master)
        dialog.grab_set()
        ttk.Label(dialog, text="为需要填色的指定项选择颜色；不选颜色的指定项保持原色。", padding=10).pack(fill="x")
        canvas = tk.Canvas(dialog, highlightthickness=0)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        content = ttk.Frame(canvas, padding=(10, 0, 10, 10))
        content.bind("<Configure>", lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        existing = parse_color_rules(self.rule_text.get(), case_sensitive=True)
        color_vars: dict[str, tk.StringVar] = {}
        for row_index, value in enumerate(values):
            ttk.Label(content, text=value).grid(row=row_index, column=0, sticky="w", pady=3)
            color_var = tk.StringVar(value=existing.get(value, ""))
            color_vars[value] = color_var
            entry = ttk.Entry(content, textvariable=color_var, width=12, state="readonly")
            entry.grid(row=row_index, column=1, sticky="w", padx=(8, 4), pady=3)

            def choose(rule_value: str = value, var: tk.StringVar = color_var) -> None:
                color = colorchooser.askcolor(color=var.get() or "#FFF2CC", title=f"选择 {rule_value} 的颜色", parent=dialog)
                if color and color[1]:
                    var.set(color[1].upper())

            def clear(var: tk.StringVar = color_var) -> None:
                var.set("")

            ttk.Button(content, text="选色", command=choose, width=8).grid(row=row_index, column=2, padx=4, pady=3)
            ttk.Button(content, text="清空", command=clear, width=8).grid(row=row_index, column=3, padx=4, pady=3)

        button_frame = ttk.Frame(dialog, padding=10)
        button_frame.pack(side="bottom", fill="x")

        def confirm() -> None:
            rules = [f"{value}={color_vars[value].get()}" for value in values if color_vars[value].get().strip()]
            self.rule_text.set(";".join(rules))
            dialog.destroy()

        ttk.Button(button_frame, text="确认", command=confirm, width=12).pack(side="right", padx=(6, 0))
        ttk.Button(button_frame, text="取消", command=dialog.destroy, width=12).pack(side="right")

    def _apply_profile_if_ready(self) -> None:
        if not self.source_headers or not self.target_headers:
            return
        source = SheetInfo(Path(self.source_path.get()), self.source_sheet.get(), self.source_header_row.get(), self.source_headers)
        target = SheetInfo(Path(self.target_path.get()), self.target_sheet.get(), self.target_header_row.get(), self.target_headers)
        profile_name = self.apply_best_profile(source, target)
        if profile_name:
            self.status.set(f"已自动套用相似配置：{profile_name}")

    def apply_best_profile(self, source: SheetInfo, target: SheetInfo) -> str | None:
        profiles = load_profiles()
        source_sig = header_signature(source.headers)
        target_sig = header_signature(target.headers)
        best_profile: dict[str, Any] | None = None
        best_score = 0.0
        for profile in profiles:
            score = (
                jaccard(source_sig, set(profile.get("source_signature", [])))
                + jaccard(target_sig, set(profile.get("target_signature", [])))
            ) / 2
            if score > best_score:
                best_score = score
                best_profile = profile
        if not best_profile or best_score < 0.72:
            return None
        source_option = option_for_header(source.headers, best_profile.get("source_header", ""))
        source_rule_option = option_for_header(source.headers, best_profile.get("source_rule_header", ""))
        target_option = option_for_header(target.headers, best_profile.get("target_header", ""))
        if not source_option or not target_option:
            return None
        self.source_column.set(source_option)
        if source_rule_option:
            self.source_rule_column.set(source_rule_option)
        self.target_column.set(target_option)
        if best_profile.get("source_start_row"):
            self.source_start_row.set(str(best_profile["source_start_row"]))
        self.source_end_row.set("" if best_profile.get("source_end_row") is None else str(best_profile["source_end_row"]))
        if best_profile.get("target_start_row"):
            self.target_start_row.set(str(best_profile["target_start_row"]))
        self.target_end_row.set("" if best_profile.get("target_end_row") is None else str(best_profile["target_end_row"]))
        self.highlight_start_column.set(
            "" if best_profile.get("highlight_start_column") is None else get_column_letter(int(best_profile["highlight_start_column"]))
        )
        self.highlight_end_column.set(
            "" if best_profile.get("highlight_end_column") is None else get_column_letter(int(best_profile["highlight_end_column"]))
        )
        self.rule_text.set(best_profile.get("rule_text", self.rule_text.get()))
        self.case_sensitive.set(bool(best_profile.get("case_sensitive", False)))
        return str(best_profile.get("name") or f"相似度 {best_score:.0%}")

    def preview(self) -> None:
        try:
            context = self._selected_context()
            case_sensitive = self.case_sensitive.get()
        except Exception as exc:
            self._show_error(exc)
            return
        self._run_task(lambda: self._preview_worker(context, case_sensitive), "正在预览，请稍候...")

    def _preview_worker(
        self,
        context: CompareContext,
        case_sensitive: bool,
    ) -> None:
        self._set_status("正在读取表 A，建立查找项与指定项规则...")
        source_sheets = [SheetInfo(path, context.source.sheet, context.source.header_row, context.source.headers) for path in context.source_paths]
        rule_map, total_rows, blank_rows = build_rule_map(
            source_sheets,
            context.source_column,
            context.source_rule_column,
            context.source_rows,
            case_sensitive=case_sensitive,
        )
        self._set_status("正在扫描表 B，按查找项匹配行...")
        matches, rows, scanned = scan_target_rule_matches(
            context.target,
            context.target_column,
            context.target_rows,
            rule_map,
            context.color_rules,
            case_sensitive=case_sensitive,
            preview_limit=PREVIEW_LIMIT,
        )
        self.latest_matches = list(matches)
        self.latest_keys_count = len(rule_map)
        self.master.after(0, lambda: self._render_preview(rows, len(matches), scanned, len(rule_map), total_rows, blank_rows))

    def _render_preview(
        self,
        rows: list[list[Any]],
        match_count: int,
        scanned: int,
        key_count: int,
        source_rows: int,
        blank_rows: int,
    ) -> None:
        self.preview_table.delete(*self.preview_table.get_children())
        for row in rows:
            values = ["" if value is None else str(value) for value in row]
            self.preview_table.insert("", "end", values=values)
        self.status.set(
            f"预览完成：表 A 读取 {source_rows} 行，去重后 {key_count} 个查找项，空值 {blank_rows} 行；"
            f"表 B 扫描 {scanned} 行，命中 {match_count} 行。"
        )

    def export(self) -> None:
        try:
            context = self._selected_context()
            case_sensitive = self.case_sensitive.get()
            target_path = Path(self.target_path.get())
            suffix = target_path.suffix or ".xlsx"
            default_name = f"{target_path.stem}_高亮结果{suffix}"
            output = filedialog.asksaveasfilename(
                title="保存高亮后的 Excel 文件",
                defaultextension=suffix,
                initialfile=default_name,
                filetypes=[("Excel 文件", f"*{suffix}"), ("所有文件", "*.*")],
            )
            if not output:
                return
        except Exception as exc:
            self._show_error(exc)
            return
        self._run_task(lambda: self._export_worker(context, Path(output), case_sensitive), "正在生成高亮文件...")

    def _export_worker(
        self,
        context: CompareContext,
        output_path: Path,
        case_sensitive: bool,
    ) -> None:
        self._set_status("正在读取表 A，建立查找项与指定项规则...")
        source_sheets = [SheetInfo(path, context.source.sheet, context.source.header_row, context.source.headers) for path in context.source_paths]
        rule_map, _, _ = build_rule_map(
            source_sheets,
            context.source_column,
            context.source_rule_column,
            context.source_rows,
            case_sensitive=case_sensitive,
        )
        self._set_status("正在扫描表 B，计算需要填色的行...")
        matches, _, scanned = scan_target_rule_matches(
            context.target,
            context.target_column,
            context.target_rows,
            rule_map,
            context.color_rules,
            case_sensitive=case_sensitive,
            preview_limit=0,
        )
        if not matches:
            self.master.after(
                0,
                lambda: messagebox.showinfo(
                    "没有命中",
                    f"表 A 已读取 {len(rule_map)} 个非空查找项；表 B 已扫描 {scanned} 行，但没有匹配到需要填色的行。\n\n"
                    "请检查：两边是否选了同一种编号/姓名列、行范围是否包含数据、单元格是否有隐藏空格或格式差异。",
                ),
            )
            return
        self._set_status("正在写入填充颜色并保存文件...")
        apply_highlight(context.target, matches, "", context.highlight_columns, output_path)
        self.save_current_profile(context, case_sensitive)
        self.master.after(
            0,
            lambda: self._export_done(output_path, len(matches), scanned),
        )

    def _export_done(self, output_path: Path, match_count: int, scanned: int) -> None:
        self.status.set(f"已生成：{output_path}；表 B 扫描 {scanned} 行，填色 {match_count} 行。")
        messagebox.showinfo("完成", f"已生成高亮文件：\n{output_path}\n\n共填色 {match_count} 行。")

    def save_current_profile(
        self,
        context: CompareContext,
        case_sensitive: bool,
    ) -> None:
        profiles = load_profiles()
        source_header = context.source.headers[context.source_column - 1] if len(context.source.headers) >= context.source_column else ""
        source_rule_header = context.source.headers[context.source_rule_column - 1] if len(context.source.headers) >= context.source_rule_column else ""
        target_header = context.target.headers[context.target_column - 1] if len(context.target.headers) >= context.target_column else ""
        profile = {
            "name": f"{source_header} → {target_header}，按 {source_rule_header} 填色",
            "source_signature": sorted(header_signature(context.source.headers)),
            "target_signature": sorted(header_signature(context.target.headers)),
            "source_header": source_header,
            "source_rule_header": source_rule_header,
            "target_header": target_header,
            "source_header_row": context.source.header_row,
            "target_header_row": context.target.header_row,
            "source_start_row": context.source_rows.start,
            "source_end_row": context.source_rows.end,
            "target_start_row": context.target_rows.start,
            "target_end_row": context.target_rows.end,
            "highlight_start_column": context.highlight_columns.start,
            "highlight_end_column": context.highlight_columns.end,
            "rule_text": self.rule_text.get(),
            "case_sensitive": case_sensitive,
        }
        profiles = [item for item in profiles if item.get("name") != profile["name"]]
        save_profiles([profile, *profiles])

    def _selected_context(self) -> CompareContext:
        source = self._sheet_info(kind="source", read_headers_now=not self.source_headers)
        target = self._sheet_info(kind="target", read_headers_now=not self.target_headers)
        source_paths = excel_files_in_source(Path(self.source_path.get()))
        source_column = parse_column_option(self.source_column.get())
        source_rule_column = parse_column_option(self.source_rule_column.get())
        target_column = parse_column_option(self.target_column.get())
        if source_column < 1 or source_rule_column < 1 or target_column < 1:
            raise ValueError("查找项列或指定项列必须大于 0。")
        source_rows = parse_row_range(
            self.source_start_row.get(),
            self.source_end_row.get(),
            source.header_row + 1,
            "表 A ",
        )
        target_rows = parse_row_range(
            self.target_start_row.get(),
            self.target_end_row.get(),
            target.header_row + 1,
            "表 B ",
        )
        highlight_columns = parse_column_range(
            self.highlight_start_column.get(),
            self.highlight_end_column.get(),
        )
        validate_source_sheets(source_paths, source.sheet, source.header_row)
        case_sensitive = self.case_sensitive.get()
        color_rules = parse_color_rules(self.rule_text.get(), case_sensitive=case_sensitive)
        if not color_rules:
            raise ValueError("请先配置至少一个指定项颜色；未配置颜色的指定项会保持原色。")
        return CompareContext(
            source=source,
            source_paths=source_paths,
            target=target,
            source_column=source_column,
            source_rule_column=source_rule_column,
            target_column=target_column,
            source_rows=source_rows,
            target_rows=target_rows,
            highlight_columns=highlight_columns,
            color_rules=color_rules,
        )

    def _sheet_info(self, *, kind: str, read_headers_now: bool) -> SheetInfo:
        if kind == "source":
            path_text = self.source_path.get()
            sheet = self.source_sheet.get()
            header_row = self.source_header_row.get()
            headers = self.source_headers
        else:
            path_text = self.target_path.get()
            sheet = self.target_sheet.get()
            header_row = self.target_header_row.get()
            headers = self.target_headers
        if not path_text:
            raise ValueError("请先选择两个 Excel 文件。")
        if not sheet:
            raise ValueError("请先选择工作表。")
        path = Path(path_text)
        if not path.exists():
            raise ValueError(f"文件不存在：{path}")
        if header_row < 1:
            raise ValueError("表头行必须大于 0。")
        if read_headers_now:
            read_path = excel_files_in_source(path)[0] if kind == "source" else path
            headers = read_headers(read_path, sheet, header_row)
            path = read_path if kind == "source" else path
        if not headers:
            raise ValueError("没有读取到表头，请确认表头行是否正确。")
        return SheetInfo(path=path, sheet=sheet, header_row=header_row, headers=headers)

    def _run_task(self, func: Any, status_text: str = "处理中，请稍候...") -> None:
        if self._busy:
            self.status.set("当前任务还在处理，请稍候。")
            return
        self._set_busy(True, status_text)

        def worker() -> None:
            try:
                func()
            except Exception as exc:
                self.master.after(0, lambda error=exc: self._show_error(error))
            finally:
                self.master.after(0, lambda: self._set_busy(False))

        threading.Thread(target=worker, daemon=True).start()

    def _set_status(self, text: str) -> None:
        self.master.after(0, lambda: self.status.set(text))

    def _set_busy(self, busy: bool, status_text: str | None = None) -> None:
        self._busy = busy
        state = "disabled" if busy else "normal"
        for button in self.task_buttons:
            button.configure(state=state)
        if status_text:
            self.status.set(status_text)

    def _show_error(self, exc: Exception) -> None:
        self.status.set(f"失败：{exc}")
        messagebox.showerror("操作失败", str(exc))

def main() -> None:
    root = tk.Tk()
    style = ttk.Style(root)
    if "vista" in style.theme_names():
        style.theme_use("vista")
    ExcelCompareApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
